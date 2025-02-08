import openpyxl
import numpy as np

#encabezados
encabezados=['Tensiones 1','tensiones 2','Tensiones 3']



# Listas de ejemplo
lista1 = [1, 2, 3, 4, 5]
lista2 = ['A', 'B', 'C', 'D', 'E']
lista3 = [10.5, 20.3, 30.1, 40.7, 50.9]


def datos2excel(encabezados,lista1,lista2,lista3):

# Crear un nuevo libro de trabajo y seleccionar la hoja activa
    libro = openpyxl.Workbook()
    hoja = libro.active

# colocar encabezados
    for i in np.arange(0,len(encabezados)):
        hoja.cell(row=1, column=i+1, value=encabezados[i])



# Escribir las listas en columnas diferentes
    for i in np.arange(0,len(lista1)):
        hoja.cell(row=i+2, column=1, value=lista1[i])
        hoja.cell(row=i+2, column=2, value=lista2[i])
        hoja.cell(row=i+2, column=3, value=lista3[i])


# Guardar el archivo Excel
    libro.save('listas_en_columnas.xlsx')

    print("Archivo Excel guardado correctamente.")




datos2excel(encabezados,lista1,lista2,lista3)