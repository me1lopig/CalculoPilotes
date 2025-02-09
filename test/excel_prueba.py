import openpyxl



# Función que acepta un número variable de listas y encabezados
def guardar_listas_en_excel(nombre_archivo, encabezados, *listas):
    # Verificar que el número de encabezados coincida con el número de listas
    if len(encabezados) != len(listas):
        raise ValueError("El número de encabezados debe coincidir con el número de listas.")

    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    libro = openpyxl.Workbook()
    hoja = libro.active

    # Escribir los encabezados en la primera fila
    for columna, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=encabezado)

    # Escribir cada lista en una columna diferente, empezando desde la segunda fila
    for columna, lista in enumerate(listas, start=1):
        for fila, valor in enumerate(lista, start=2):  # Empezar desde la fila 2
            hoja.cell(row=fila, column=columna, value=valor)

    # Guardar el archivo Excel
    libro.save(nombre_archivo)
    print(f"Archivo Excel '{nombre_archivo}' guardado correctamente.")

# Ejemplo de uso
lista1 = [1, 2, 3, 4, 5]
lista2 = ['A', 'B', 'C', 'D', 'E']
lista3 = ['001', '002', '003', '004', '005']
lista4 = [10.5, 20.3, 30.1, 40.7, 50.9]

# Lista de encabezados
encabezados = ["Números", "Letras", "Códigos", "Decimales"]

# Llamar a la función con encabezados y listas
guardar_listas_en_excel('listas_en_columnas.xlsx', encabezados, lista1, lista2, lista3, lista4)