
# Libreria de funciones usadas en los cálculos 

# Listado de funciones (actualizar según se añaden funciones)

    # crea_directorio(), genera el directorio de trabajo para alojar los archivos de proyecto

    # datos_pilotes, importa los datos de los pilotes a calcular
    # datos_terreno, importa de una hoja excel los datos del terreno
    # reemplaza_none, elimina los valores None de la toma de datos y los sustituye por un numero (0)
 

    # parametro_terreno, obtiene cualquier parámetro del terreno en función de la profundidad
    # n_freatico, calcula si a una profundidad existe nivel freático

    # insertar_valor inserta un valor en una lista y la ordena con el nuevo valor es función auxiliar de presion_total
    # obtener_maximo_menor calculamos el valor de la cota inferior del estrato superior más cercano a la cota que se introduce
        #  es función  auxiliar de presion_total
    # insertar_valor, inserta un valor de forma ordenada dentro de una lista

    # presion_total, calcula la presión total en un punto del terreno
    # presion media, calcula la presión media de un tramo de fuste de pilote

    # promedioPunta, cálculo del promedio de un parametro en la zona de punta 3D abajo y 6D hacia arriba
    # grafica_tensiones, calcula las tensiones en el intervalo de datos del archivo datos_terreno y da la grafica y la tabla

    # grafico_grupo, plotea varias magnitudes  o de forma individual

    # qp_CTE_gr, calcula la resistencia en punta en el caso de situación drenada
    # qp_CTE_cohesivos, calcula la resistencia en punta en el caso de situación no drenada

    # tf_CTE_gr, calcula la resistencia en fuste en situación drenada
    # tf_CTE_chesivos, calcula la resistencia en fuste en situación no drenada



# librerias 
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from docx import Document
from docx.shared import Cm, Pt
import os
from datetime import datetime


# Grupo de funciones


def crea_directorio():
    # creación del directorio de trabajo para guardar resultados
    now = datetime.now()
    directorio=(str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second))
    os.mkdir(directorio)
    return directorio


def reemplaza_None(lista, numero):
    # coloca un numero (0 en este caso o cualquier otro) en los datos faltantes en la hoja excel de datos (None)
    return [numero if x is None else x for x in lista]


def datos_terreno(archivo):

    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active

    # importacion de variables del terreno
    
    # geometría de las capas
    espesor=[]
    cotas=[]
    # valores físicos
    pe_seco=[]
    pe_saturado=[]
    
    # valores de resistencia
    cu=[]
    cohesion=[]
    fi=[]

    # tipo de datos (encabezados)
    tipo_datos=[]

    # tipo de calculo a realizar, drenado o no drenado
    tipo_calculo=[]

    # Toma de datos de iteración

    for row in hoja.iter_rows():
        espesor.append(row[0].value)
        espesor[0]=0
        az=sum(espesor) # vector de espesores
 
        pe_seco.append(row[2].value)
        pe_seco[0]=0
        pe_saturado.append(row[3].value)
        pe_saturado[0]=0
    

        cu.append(row[4].value)
        cu[0]=0

        cohesion.append(row[5].value)
        cohesion[0]=0

        fi.append(row[6].value)
        fi[0]=0

        tipo_calculo.append(row[7].value)
        tipo_calculo[0]=''



    nivel_freatico=hoja.cell(row=2, column=2).value

    for i in np.arange(len(espesor)):
        cotas.append(sum(espesor[0:i+1]))
    
    # se toman los datos de la cabecera
    for filas in hoja.iter_cols():
        tipo_datos.append(filas[0].value)


    # Reemplazo de valores vacios por un 0
    cu=reemplaza_None(cu, 0)
    cohesion=reemplaza_None(cohesion, 0)
    fi=reemplaza_None(fi, 0)
    
    return espesor,cotas,az,nivel_freatico,pe_seco,pe_saturado,cu,cohesion,fi,tipo_datos,tipo_calculo


def datos_pilotes(archivo):
    # importacion de los datos de defición de los pilotes

    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active

        # importacion de variables del terreno
    
    # geometría de las capas
    diametros=[]

    for row in hoja.iter_rows():
        diametros.append(row[0].value)
 
    diametros.pop(0) # Se elimina el primer elemento
    diametros=np.divide(diametros,1000) # pasamos a metros

    Lmin=hoja.cell(row=2, column=2).value
    Lincr=hoja.cell(row=2, column=3).value
    fp=hoja.cell(row=2, column=4).value
    kr=hoja.cell(row=2, column=5).value
    f=hoja.cell(row=2, column=6).value


    return diametros,Lmin,Lincr,fp,kr,f



def parametro_terreno(cotas,zt):
    # esta funcion localiza un  parámetro del terreno para una z determinada
    # se parte de los valores discretos tomados de la excel de datos_terreno.xlsx
    # sirve para cualquier parámetro del terreno
    # ejemplo pe_saturado[ft.parametro_terreno(cotas,6)]

    for z in np.arange(len(cotas)-1):
        cota_superior=cotas[z]
        cota_inferior=cotas[z+1]
        if zt>=cota_superior and zt<=cota_inferior:
            break
    return z+1



def n_freatico(nivel_freatico,z):
    # función para calcular el nivel freático de forma continua
    # indica si un valor del terreno está o no bajo el nf
    # devuelve el valor de la altura de columna o un 0 según se esté o no en la zona del nivel freático
    if z>=nivel_freatico:
        return (z-nivel_freatico)
    else:
        return 0


# funciones auxiliares para el cálculo de la presion total

def insertar_valor(lista, valor):
    # funcion para insertar la cota en la lista de cotas y ordenarla
    lista_mod=[] 
    lista_mod=lista.copy() # se evita alterar la lista original
    lista_mod.append(valor)  # Agrega el valor a la lista
    lista_mod.sort()         # Ordena la lista en orden ascendente
    return lista_mod



def obtener_maximo_menor(lista, valor):
    # calculamos el valor de la cota inferior del estrato superior 
    # más cercano a la cota que se introduce
    maximo_menor = max(filter(lambda x: x <= valor, lista))
    return maximo_menor



def presion_total(cotas,valor_nf,pe_saturado,pe_seco,valor_cota):
    
    # cotas es la lista de cotas
    # pe_saturado, lista de pesos específicos saturados
    # pe_seco.lista de pesos específicos secos
    # nf es la profundidad del nivel freático
    # valor_cota es la profundidad a la que se va a calcular la presión total
    
    
    lista_cotas = cotas.copy() # se copia la lista de las cotas para no alterarla

    #lista ordenada contiene la lista de las cotas y la cota hasta la que se quiere calcular las tensiones
    # en caso de que ya esté el valor no se introduce de nuevo

    if valor_nf not in lista_cotas: 
        lista_valores = insertar_valor(lista_cotas, valor_nf)
    else:
        lista_valores=lista_cotas

    
    # resultado es el valor de la cota inmediatamente anterior a la que queremos calcular
    resultado = obtener_maximo_menor(lista_valores, valor_cota)
    
    
    # calculo del estrato del fondo
    peso_saturado=pe_saturado[parametro_terreno(cotas,valor_cota)] # peso especifico saturado
    peso_seco=pe_seco[parametro_terreno(cotas, valor_cota)] # peso especifico seco o aparente 
    
    # cálculo de las presiones totales y efectivas, inicio de valores
    peso=peso_seco if valor_cota<=valor_nf else peso_saturado
    presion_total=(valor_cota-resultado)*peso
    
    # resto de estratos
    for j in range(lista_valores.index(resultado),0,-1):
        espesor=lista_valores[j]-lista_valores[j-1] # espesor del estrato
        peso_saturado=pe_saturado[parametro_terreno(cotas,lista_valores[j])] # peso especifico saturado
        peso_seco=pe_seco[parametro_terreno(cotas,lista_valores[j])] # peso especifico seco
        posicion=lista_valores[j] # indica la cota del final del nivel
    
        # sumatoria de las presiones
        if (posicion<=valor_nf):
            peso=peso_seco
        else:
            peso=peso_saturado
    
        presion_total=presion_total+espesor*peso
    
    return presion_total



def grafica_tensiones(cotas,pe_seco,pe_saturado,nivel_freatico,directorio):
    valor_z=np.arange(0,max(cotas)+0.10,0.10)
    presionEfectiva=[]
    presionTotal=[]
    presionPoro=[]

    for z in valor_z:
        pe_sat=pe_saturado[parametro_terreno(cotas,z)]
        pe_ap=pe_seco[parametro_terreno(cotas,z)]
        presion=presion_total(cotas,nivel_freatico,pe_saturado,pe_seco,z)
        u_z=n_freatico(nivel_freatico,z)*9.81
        presion_efectiva=presion-u_z
   
        presionEfectiva.append(presion_efectiva)
        presionTotal.append(presion)
        presionPoro.append(u_z)
        #print(z,presion,presion_efectiva,u_z)

    presionE = (presionEfectiva, valor_z, 'darkred', 'Presion Efectiva')
    presionT = (presionTotal, valor_z, 'blue', 'Presion Total')
    presionP = (presionPoro, valor_z, 'green', 'Presion de Poro')

    # todas las tensiones
    lista_datos = [presionE, presionT, presionP]
    grafico_grupo(lista_datos, "Tensiones en el terreno",'kN/m2',directorio)



def grafico_grupo(lista_datos, titulo,etiqueta_x,directorio):
    fig, ax = plt.subplots()

    for datos in lista_datos:
        xcoord, ycoord, color, label = datos
        ax.plot(xcoord, np.multiply(ycoord, 1), marker='', color=color, linestyle='-', label=label)

    ax.invert_yaxis()
    ax.set_aspect('auto', adjustable='box')
    ax.set_xlabel(etiqueta_x)
    ax.set_ylabel("Profundidad [m]")
    ax.set_title(titulo)
    ax.legend()
    #ax.set_aspect('equal', adjustable='box')
    plt.savefig(directorio+'/'+titulo+".png") # guardado de la imagen
    #plt.show()


def promedioPunta(D,L,cotas,parametro):
    # cálculo del promedio en punta de los parámtros resistentes
    #D= diámetro pilote
    #L=9.32  longitud pilote
    # cotas tabla de cotas del terreno
    # parametro, parametro al que calcular el promedio

    incr=0.05 # incremento de paso
    
    # valores por debajo de la punta del pilote 3D
    p_suma=0
    contador=0
    for z in np.arange(L,L+3*D+incr,incr):
        p=parametro[parametro_terreno(cotas,z)]
        p_suma=p+p_suma
        contador=contador+1

    promedio3D=(p_suma)/(contador)


    # valores por encima de la punta 6D
    p_suma=0
    contador=0
    # valores por debajo de la punta del pilote
    for z in np.arange(L,L-6*D-incr,-incr):
        p=parametro[parametro_terreno(cotas,z)]
        p_suma=p+p_suma
        contador=contador+1
    
    promedio6D=(p_suma)/(contador)

    # cálculo del promedio
    promedio=(3*promedio3D+6*promedio6D)/9

    return promedio


def qp_CTE_gr(cotas,valor_nf,pe_saturado,pe_seco,fi,D,L,fp):
    # calculo de la tensión unitaria por punta según el CTE suelos granulares
    # fi, ángulo de rozamiento
    # D, diámetro del pilote
    # L longitud del pilote
    # fp factor que depende del proceso constructivo fp=2 hincados y fp=3 insitu

    # cálculo de la presion efectiva en la zona de punta

    presionTotal=presion_total(cotas,valor_nf,pe_saturado,pe_seco,L) # total
    u_z=n_freatico(valor_nf,L)*9.81 # presion de poro
    presionEfectiva=presionTotal-u_z # presion efectiva

  
    # valor promediado de fi
    fi_promedio=promedioPunta(D,L,cotas,fi)
    fi_promedio=np.deg2rad(fi_promedio) # paso a radianes

    # factor de capacidad de carga
    if fi_promedio==0:
        Nq=0
    else:
        Nq=(1+np.sin(fi_promedio))/(1-np.sin(fi_promedio))*np.exp(np.pi*np.tan(fi_promedio))

    # Cálculo de las tensiones unitarias 
    qp=fp*presionEfectiva*Nq # pilotes en KPa

    # limitacion de los 20 MPa
    if (qp/1000>20):
        qp=20000

    # calculo de la carga de hundimiento y admisible
    area=0.25*np.pi*D**2
    Qhp=qp*area

    return qp,Qhp


def qp_CTE_cohesivos(cotas,cu,D,L):
    # calculo de la tensión unitaria por punta según el CTE suelos cohesivos
    # cu, cohesion sin drenaje
    # L longitud del pilote

  
    # valor promediado de cu
    cu_promedio=promedioPunta(D,L,cotas,cu)

    # factor de capacidad de carga
    Np=9

    # Cálculo de las tensiones unitarias 
    qp=cu_promedio*Np # pilotes en KPa


    # calculo de la carga de hundimiento y admisible
    area=0.25*np.pi*D**2
    Qhp=qp*area

    return qp,Qhp


def tf_CTE_gr(cotas,nivel_freatico,pe_seco,pe_saturado,fi,D,L,kr,f,tipo_calculo):

    # calculo de la tensión unitaria por punta según el CTE suelos granulares
    # fi, ángulo de rozamiento
    # D, diámetro del pilote
    # L longitud del pilote
    # f factor que depende del proceso constructivo y tipo de material 
        # f=1 insitu hormigón y madera
        # f=0.9 prefabricados hormigón y madera
        # f=0.8 de acero
    # kr factor que depende del proceso constructivo 
        # kr=1 hncados  
        # kr=0.75 perforados

    # Cálculo de las tensiones efectivas medias en el nivel considerado

    def presionEfectivaMedia(zfinf,zfsup):
    # funcion auxiliar para el cálculo de la tension efectiva media 
    
        presion=presion_total(cotas,nivel_freatico,pe_saturado,pe_seco,zfinf)
        u_z=n_freatico(nivel_freatico,zfinf)*9.81
        presion_efectiva_inf=presion-u_z

    
        presion=presion_total(cotas,nivel_freatico,pe_saturado,pe_seco,zfsup)
        u_z=n_freatico(nivel_freatico,zfsup)*9.81
        presion_efectiva_sup=presion-u_z

        presion_media=(presion_efectiva_inf+presion_efectiva_sup)*0.5

        return presion_media


    if nivel_freatico not in cotas: 
        listafuste = insertar_valor(cotas, nivel_freatico)
    else:
        listafuste=cotas


    # listados de resultados
    listaTensionesFuste=[]
    listaLongitudesFuste=[]
    listaTensionesUnitarias=[]
    ListaCargaHundimiento=[]

    for z in np.arange(0,len(listafuste)-1):

        zfinf=listafuste[z]
        zfsup=listafuste[z+1]


        if (zfinf<L) and (zfsup<L):
            presion_media=presionEfectivaMedia(zfinf,zfsup)
            listaTensionesFuste.append(presion_media)
            listaLongitudesFuste.append(zfsup-zfinf)
        
        elif (zfinf<L) and (L<=zfsup):
        
            presion_media=presionEfectivaMedia(zfinf,L)
            listaTensionesFuste.append(presion_media)
            listaLongitudesFuste.append(L-zfinf)  
            break # se llega al tope maximo


    listaLongitudesFusteAcumulada=np.cumsum(listaLongitudesFuste)

    for t in np.arange(0,len(listaTensionesFuste)):
        fi_deg=fi[parametro_terreno(cotas,listaLongitudesFusteAcumulada[t])]
        fi_rad=float(np.deg2rad(fi_deg))


        tf=listaTensionesFuste[t]*kr*f*np.tan(fi_rad)
        #limitacion de 120 kPA
        if tf>120:
            tf=120
        Qhfi=tf*np.pi*D*listaLongitudesFuste[t] # cargas de hundimieto parciales
        listaTensionesUnitarias.append(tf) # creacion de la lista de las tensiones unitarias
        ListaCargaHundimiento.append(Qhfi)

    return listaTensionesUnitarias,ListaCargaHundimiento,listaLongitudesFusteAcumulada




def tf_CTE_cohesivos(cotas,cu,D,L):

    # calculo de la tensión unitaria por punta según el CTE suelos granulares
    # cu, cohesion sin drnaje
    # D, diámetro del pilote
    # L longitud del pilote

    listafuste=cotas


    # listados de resultados
    listaCohesiones=[]
    listaLongitudesFuste=[]
    listaTensionesUnitarias=[]
    ListaCargaHundimiento=[]

    for z in np.arange(0,len(listafuste)-1):

        zfinf=listafuste[z]
        zfsup=listafuste[z+1]


        if (zfinf<L) and (zfsup<L):
            cohesion=cu[parametro_terreno(cotas,(zfsup+zfinf)*0.5)]
            listaCohesiones.append(cohesion)
            listaLongitudesFuste.append(zfsup-zfinf)
        
        elif (zfinf<L) and (L<=zfsup):
        
            cohesion=cu[parametro_terreno(cotas,(zfsup+zfinf)*0.5)]
            listaCohesiones.append(cohesion)
            listaLongitudesFuste.append(L-zfinf)  
            break # se llega al tope maximo

    listaLongitudesFusteAcumulada=np.cumsum(listaLongitudesFuste)

    for t in np.arange(0,len(listaCohesiones)):
 
        tf=(listaCohesiones[t]*100)/(100+(listaCohesiones[t]))
        Qhfi=tf*np.pi*D*listaLongitudesFuste[t] # cargas de hundimieto parciales
        listaTensionesUnitarias.append(tf) # creacion de la lista de las tensiones unitarias
        ListaCargaHundimiento.append(Qhfi)



    return listaTensionesUnitarias,ListaCargaHundimiento,listaLongitudesFusteAcumulada




def cargaHundimientoFuste(ListaLongitudesFusteAcumuladasGr,ListaLongitudesFusteAcumuladasCo,ListaCargaHundimientoGr,ListaCargaHundimientoCo,cotas,tipo_calculo):
    
    # seleccion del tipo de cargas por fuste
    Qhf=0

    # seleccion de las cargas por fuste situacion drenada
    for tipo in np.arange(0,len(ListaLongitudesFusteAcumuladasGr)):
        zt=ListaLongitudesFusteAcumuladasGr[tipo]
        tipoCalculo=tipo_calculo[parametro_terreno(cotas,zt)]
        if tipoCalculo=='d':
            Qhf+=ListaCargaHundimientoGr[tipo]


    # seleccion de las cargas por fuste situacion nodrenada
    for tipo in np.arange(0,len(ListaLongitudesFusteAcumuladasCo)):
        zt=ListaLongitudesFusteAcumuladasCo[tipo]
        tipoCalculo=tipo_calculo[parametro_terreno(cotas,zt)]
        if tipoCalculo=='nd':
            Qhf+=ListaCargaHundimientoCo[tipo]


    return Qhf

