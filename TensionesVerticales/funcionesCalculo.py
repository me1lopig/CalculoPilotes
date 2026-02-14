# Libreria de funciones usadas en los cálculos de tensiones del terreno

import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import os
from datetime import datetime

# --- FUNCIONES DE GESTIÓN Y DATOS ---

def crea_directorio():
    # creación del directorio de trabajo para guardar resultados
    now = datetime.now()
    directorio = (str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second))
    # Comprobación simple para evitar errores si se ejecuta muy rápido
    if not os.path.exists(directorio):
        os.mkdir(directorio)
    return directorio

def datos_terreno(archivo):
    libro = openpyxl.load_workbook(archivo)
    hoja = libro.active

    # importacion de variables del terreno
    espesor = []
    cotas = []
    pe_seco = []
    pe_saturado = []

    # Toma de datos de iteración (Solo columnas necesarias: 0, 2 y 3)
    for row in hoja.iter_rows():
        # Geometría (Columna A - índice 0)
        valor_espesor = row[0].value
        espesor.append(valor_espesor if valor_espesor is not None else 0)
        
        # Peso Seco (Columna C - índice 2)
        valor_seco = row[2].value
        pe_seco.append(valor_seco if valor_seco is not None else 0)
        
        # Peso Saturado (Columna D - índice 3)
        valor_sat = row[3].value
        pe_saturado.append(valor_sat if valor_sat is not None else 0)
    
    # Corrección del elemento 0 (fila de encabezados o valor nulo inicial)
    espesor[0] = 0
    pe_seco[0] = 0
    pe_saturado[0] = 0

    # Nivel freático (Celda B2)
    nivel_freatico = hoja.cell(row=2, column=2).value

    # Calculo de cotas acumuladas
    for i in np.arange(len(espesor)):
        cotas.append(sum(espesor[0:i+1]))
    
    # Solo devolvemos lo que usa el cálculo de tensiones
    return cotas, nivel_freatico, pe_seco, pe_saturado

# --- FUNCIONES DE CÁLCULO DE TENSIONES ---

def parametro_terreno(cotas, zt):
    # localiza el índice del estrato para una profundidad zt
    # Se optimiza el bucle para devolver el índice correcto
    indice = 1
    for z in range(len(cotas)-1):
        cota_superior = cotas[z]
        cota_inferior = cotas[z+1]
        if zt >= cota_superior and zt <= cota_inferior:
            indice = z + 1
            break
    return indice

def n_freatico(nivel_freatico, z):
    # calcula la altura de la columna de agua
    if z >= nivel_freatico:
        return (z - nivel_freatico)
    else:
        return 0

def insertar_valor(lista, valor):
    # funcion auxiliar para presion_total
    lista_mod = lista.copy() 
    if valor not in lista_mod:
        lista_mod.append(valor)  
    lista_mod.sort()         
    return lista_mod

def obtener_maximo_menor(lista, valor):
    # funcion auxiliar para presion_total
    # Devuelve el valor más grande de la lista que sea menor o igual al valor dado
    filtro = [x for x in lista if x <= valor]
    if not filtro:
        return 0
    return max(filtro)

def presion_total(cotas, valor_nf, pe_saturado, pe_seco, valor_cota):
    
    lista_cotas = cotas.copy() 

    # Insertamos el nivel freático como una cota virtual si no existe
    if valor_nf not in lista_cotas: 
        lista_valores = insertar_valor(lista_cotas, valor_nf)
    else:
        lista_valores = lista_cotas
    
    # Cota inferior del estrato o tramo inmediatamente superior
    resultado = obtener_maximo_menor(lista_valores, valor_cota)
    
    # Parámetros del punto actual
    idx_terreno = parametro_terreno(cotas, valor_cota)
    peso_saturado_val = pe_saturado[idx_terreno] 
    peso_seco_val = pe_seco[idx_terreno] 
    
    # Peso del último tramo (desde 'resultado' hasta 'valor_cota')
    peso = peso_seco_val if valor_cota <= valor_nf else peso_saturado_val
    presion_total_calc = (valor_cota - resultado) * peso
    
    # Suma de presiones de los estratos superiores
    # Iteramos hacia atrás desde el tramo anterior
    idx_resultado = lista_valores.index(resultado)
    
    for j in range(idx_resultado, 0, -1):
        cota_actual = lista_valores[j]
        cota_anterior = lista_valores[j-1]
        espesor = cota_actual - cota_anterior
        
        # Para saber las propiedades, usamos el punto medio del tramo
        cota_media = (cota_actual + cota_anterior) / 2
        idx_t = parametro_terreno(cotas, cota_media)
        
        p_sat = pe_saturado[idx_t]
        p_seco = pe_seco[idx_t]
        
        # Determinamos si el tramo está bajo agua o no
        # Usamos la cota superior del tramo (cota_actual) para decidir
        if cota_actual <= valor_nf:
            peso_tramo = p_seco
        else:
            peso_tramo = p_sat
    
        presion_total_calc += espesor * peso_tramo
    
    return presion_total_calc

# --- FUNCIONES DE GRÁFICOS Y SALIDA ---

def grafica_tensiones(cotas, pe_seco, pe_saturado, nivel_freatico, directorio):
    valor_z = np.arange(0, max(cotas) + 0.10, 0.10)
    presionEfectiva = []
    presionTotal = []
    presionPoro = []

    for z in valor_z:
        presion = presion_total(cotas, nivel_freatico, pe_saturado, pe_seco, z)
        u_z = n_freatico(nivel_freatico, z) * 9.81
        presion_efectiva = presion - u_z
   
        presionEfectiva.append(presion_efectiva)
        presionTotal.append(presion)
        presionPoro.append(u_z)

    presionE = (presionEfectiva, valor_z, 'darkred', 'Presion Efectiva')
    presionT = (presionTotal, valor_z, 'blue', 'Presion Total')
    presionP = (presionPoro, valor_z, 'green', 'Presion de Poro')

    lista_datos = [presionE, presionT, presionP]
    grafico_grupo(lista_datos, "Tensiones en el terreno", 'Tensión [kN/m2]', directorio)

    # creacion de la tabla de datos en excel
    encabezados = ['Profundidad (m)', 'Presion de Poro (kPa)', 'Presion Efectiva (kPa)', 'Presion Total (kPa)']
    guardar_listas_en_excel(os.path.join(directorio, 'TensionesTerreno.xlsx'), encabezados, valor_z, presionPoro, presionEfectiva, presionTotal)

def grafico_grupo(lista_datos, titulo, etiqueta_x, directorio):
    fig, ax = plt.subplots(figsize=(8, 10))

    for datos in lista_datos:
        xcoord, ycoord, color, label = datos
        ax.plot(xcoord, ycoord, color=color, linestyle='-', label=label)

    ax.invert_yaxis()
    ax.set_xlabel(etiqueta_x)
    ax.set_ylabel("Profundidad [m]")
    ax.set_title(titulo)
    ax.legend()
    ax.grid(True, linestyle='--', alpha=0.7)
    
    ruta_imagen = os.path.join(directorio, titulo + ".png")
    plt.savefig(ruta_imagen)
    plt.close() # Importante cerrar la figura para liberar memoria

def guardar_listas_en_excel(nombre_archivo, encabezados, *listas):
    if len(encabezados) != len(listas):
        raise ValueError("El número de encabezados debe coincidir con el número de listas.")

    libro = openpyxl.Workbook()
    hoja = libro.active

    for columna, encabezado in enumerate(encabezados, start=1):
        hoja.cell(row=1, column=columna, value=encabezado)

    for columna, lista in enumerate(listas, start=1):
        for fila, valor in enumerate(lista, start=2): 
            hoja.cell(row=fila, column=columna, value=valor)

    libro.save(nombre_archivo)
    print(f"Archivo Excel guardado en: {nombre_archivo}")