import os
import zipfile
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def _aplicar_estilo_hoja(ws, headers, titulo='Cálculos Hiley'):
    thin_side = Side(style='thin', color='B0B0B0')
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    fill_title = PatternFill('solid', fgColor='1F4E79')
    fill_header = PatternFill('solid', fgColor='D9E1F2')

    font_title = Font(bold=True, color='FFFFFF', size=14)
    font_header = Font(bold=True, color='1F1F1F')

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell_title = ws.cell(row=1, column=1, value=titulo)
    cell_title.fill = fill_title
    cell_title.font = font_title
    cell_title.alignment = align_center

    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center
        c.border = border_thin

    col_widths = {1: 22, 2: 10, 3: 10, 4: 8, 5: 10, 6: 10, 7: 10, 8: 10, 9: 24, 10: 24}
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = 'B3'

    return border_thin, align_left, align_center


def _escribir_df_en_hoja(ws, df_calc, headers, titulo):
    border_thin, align_left, align_center = _aplicar_estilo_hoja(ws, headers, titulo=titulo)

    start_row = 3
    for i in range(len(df_calc)):
        row_vals = df_calc.iloc[i].tolist()
        for j, val in enumerate(row_vals, start=1):
            c = ws.cell(row=start_row + i, column=j, value=val)
            c.border = border_thin
            if j == 1:
                c.alignment = align_left
            else:
                c.alignment = align_center
                if isinstance(val, (int, float, np.integer, np.floating)):
                    c.number_format = '0.00'


def generar_hiley(
    ruta_dprg='DPRG.XLSX',
    salida_xlsx='Hiley_calculos.xlsx',
    carpeta_imagenes='imagenes_png',
    salida_zip='Entrega_imagenes_y_calculos.zip'
):
    """Genera un Excel de cálculos Hiley  y PNGs por ensayo.

    - En el Excel crea una pestaña por ensayo detectado (P-01, P-02, etc.)
    - Además genera un ZIP con el Excel y las imágenes.

    Entradas esperadas en DPRG:
    - Descripción Muestra
    - Profundidad
    - Número de Golpes
    """

    df_dprg = pd.read_excel(ruta_dprg, sheet_name=0)

    col_desc = 'Descripción Muestra'
    col_depth = 'Profundidad'
    col_blows = 'Número de Golpes'

    df_base = df_dprg[[col_desc, col_depth, col_blows]].copy()
    df_base[col_depth] = pd.to_numeric(df_base[col_depth], errors='coerce')
    df_base[col_blows] = pd.to_numeric(df_base[col_blows], errors='coerce')
    df_base = df_base.dropna().sort_values([col_desc, col_depth]).reset_index(drop=True)

    headers = [
        col_desc,
        'Z(m)',
        'N(20)',
        'F',
        'a',
        'e',
        'n',
        'c',
        'Presión Característica (kg/cm2)',
        'Presión Admisible (kg/cm2)'
    ]

    z_vals = df_base[col_depth].astype(float).to_numpy()
    n20_vals = df_base[col_blows].astype(float).to_numpy()

    a_vals = z_vals / 10.0 + 0.25
    e_vals = 20.0 / n20_vals
    n_vals = 0.7 - 0.7 / 19.0 * (e_vals - 1.0)
    c_vals = 0.5 - 0.5 / 19.0 * (e_vals - 1.0)

    pres_car = 63.5 * 76.0 * (1.0 + (n_vals ** 2) * a_vals) / ((e_vals + c_vals) * (1.0 + a_vals) * 20.0)
    pres_adm = pres_car / 50.0

    df_calc_all = pd.DataFrame({
        headers[0]: df_base[col_desc].astype(str).to_numpy(),
        headers[1]: z_vals,
        headers[2]: n20_vals,
        headers[3]: np.full_like(z_vals, 50.0, dtype=float),
        headers[4]: a_vals,
        headers[5]: e_vals,
        headers[6]: n_vals,
        headers[7]: c_vals,
        headers[8]: pres_car,
        headers[9]: pres_adm
    })

    os.makedirs(carpeta_imagenes, exist_ok=True)
    ensayos = sorted(df_calc_all[col_desc].dropna().unique().tolist())
    rutas_png = []

    for ensayo in tqdm(ensayos, desc='Generando PNGs'):
        df_e = df_calc_all[df_calc_all[col_desc] == ensayo].sort_values('Z(m)')

        ruta_pa = os.path.join(carpeta_imagenes, 'Pa_vs_Z_' + str(ensayo) + '.png')
        plt.figure(figsize=(6.5, 5.5))
        plt.plot(df_e['Presión Admisible (kg/cm2)'], df_e['Z(m)'], marker='o', linewidth=2)
        plt.gca().invert_yaxis()
        plt.grid(True, alpha=0.3)
        plt.xlabel('Presión Admisible (kg/cm2)')
        plt.ylabel('Z(m)')
        plt.title('Ensayo ' + str(ensayo) + '  Presión Admisible vs Profundidad')
        plt.tight_layout()
        plt.savefig(ruta_pa, dpi=200)
        plt.close()
        rutas_png.append(ruta_pa)

        ruta_n = os.path.join(carpeta_imagenes, 'Golpeos_vs_Z_' + str(ensayo) + '.png')
        plt.figure(figsize=(6.5, 5.5))
        plt.plot(df_e['N(20)'], df_e['Z(m)'], marker='o', linewidth=2, color='#8B0000')
        plt.gca().invert_yaxis()
        plt.grid(True, alpha=0.3)
        plt.xlabel('Número de Golpeos  N(20)')
        plt.ylabel('Z(m)')
        plt.title('Ensayo ' + str(ensayo) + '  Golpeos vs Profundidad')
        plt.tight_layout()
        plt.savefig(ruta_n, dpi=200)
        plt.close()
        rutas_png.append(ruta_n)

    wb = Workbook()

    # Eliminar la hoja por defecto y crear una por ensayo
    ws_def = wb.active
    wb.remove(ws_def)

    for ensayo in ensayos:
        sheet_name = str(ensayo)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws_e = wb.create_sheet(title=sheet_name)
        df_e = df_calc_all[df_calc_all[col_desc] == ensayo].sort_values('Z(m)')
        _escribir_df_en_hoja(ws_e, df_e, headers, titulo='Cálculos Hiley ' + str(ensayo))

    wb.save(salida_xlsx)

    with zipfile.ZipFile(salida_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.write(salida_xlsx, arcname=os.path.basename(salida_xlsx))
        for p in rutas_png:
            zf.write(p, arcname=os.path.join(os.path.basename(carpeta_imagenes), os.path.basename(p)))

    return salida_xlsx, rutas_png, salida_zip


if __name__ == '__main__':
    xlsx_path, png_list, zip_path = generar_hiley()
    print(xlsx_path)
    print(zip_path)
