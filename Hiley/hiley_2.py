import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generar_hiley(
    ruta_dprg='DPRG.XLSX',
    salida_xlsx='HILEY.xlsx'
):
    # Leer DPRG
    df_dprg = pd.read_excel(ruta_dprg, sheet_name=0)

    # Columnas esperadas en DPRG
    col_desc = 'Descripción Muestra'
    col_depth = 'Profundidad'
    col_blows = 'Número de Golpes'

    # Limpiar y preparar
    df_base = df_dprg[[col_desc, col_depth, col_blows]].copy()
    df_base[col_depth] = pd.to_numeric(df_base[col_depth], errors='coerce')
    df_base[col_blows] = pd.to_numeric(df_base[col_blows], errors='coerce')
    df_base = df_base.dropna().sort_values([col_desc, col_depth]).reset_index(drop=True)

    # Encabezados finales en el Excel
    headers = [
        'Z(m)',
        'N(20)',
        'F',
        'a',
        'e',
        'n',
        'c',
        'Presión Característica (kg/cm2)',
        'Presión Admisible (kg/cm2)'
    ]

    # Cálculo Hiley en Python (equivalente a las fórmulas extraídas de HILEY.xlsx)
    def calcular_hiley(df_in):
        z_vals = df_in[col_depth].astype(float).to_numpy()
        n20_vals = df_in[col_blows].astype(float).to_numpy()

        a_vals = z_vals / 10.0 + 0.25
        e_vals = 20.0 / n20_vals
        n_vals = 0.7 - 0.7 / 19.0 * (e_vals - 1.0)
        c_vals = 0.5 - 0.5 / 19.0 * (e_vals - 1.0)

        pres_car = 63.5 * 76.0 * (1.0 + (n_vals ** 2) * a_vals) / ((e_vals + c_vals) * (1.0 + a_vals) * 20.0)
        pres_adm = pres_car / 50.0

        df_out = pd.DataFrame({
            headers[0]: z_vals,
            headers[1]: n20_vals,
            headers[2]: np.full_like(z_vals, 50.0, dtype=float),
            headers[3]: a_vals,
            headers[4]: e_vals,
            headers[5]: n_vals,
            headers[6]: c_vals,
            headers[7]: pres_car,
            headers[8]: pres_adm
        })
        return df_out

    # Estilos (elegante, sin tabla ni filtros)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    thin_side = Side(style='thin', color='B0B0B0')
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    fill_title = PatternFill('solid', fgColor='1F4E79')
    fill_header = PatternFill('solid', fgColor='D9E1F2')

    font_title = Font(bold=True, color='FFFFFF', size=14)
    font_header = Font(bold=True, color='1F1F1F')

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def formatear_hoja(ws, desc_name, n_rows):
        # Fila 1: título
        ws.merge_cells('A1:I1')
        ws['A1'] = 'Cálculo Hiley - ' + str(desc_name) + ' (valores fijos)'
        ws['A1'].font = font_title
        ws['A1'].fill = fill_title
        ws['A1'].alignment = align_left
        ws.row_dimensions[1].height = 26

        # Fila 2: encabezados (sin filtros)
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=j, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin
        ws.row_dimensions[2].height = 32

        # Formatos de número y bordes
        start_row = 3
        end_row = start_row + n_rows - 1

        for r in range(start_row, end_row + 1):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.border = border_thin
                cell.alignment = align_center

                if c == 1:
                    cell.number_format = '0.0'     # Z(m)
                elif c in [2, 3]:
                    cell.number_format = '0'       # N(20), F
                elif c in [8, 9]:
                    cell.number_format = '0.00'    # Presiones (2 decimales)
                else:
                    cell.number_format = '0.0000'  # a, e, n, c

        # Anchos de columna
        widths = [10, 10, 6, 8, 8, 8, 8, 30, 28]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # Congelar encabezado
        ws.freeze_panes = 'A3'

        # Configuración básica de impresión
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = '1:2'

    # Crear una hoja por cada Descripción Muestra
    for desc_val, df_grp in df_base.groupby(col_desc):
        df_calc = calcular_hiley(df_grp)

        sheet_name = str(desc_val)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        base_name = sheet_name
        k = 1
        while sheet_name in wb_out.sheetnames:
            suffix = '_' + str(k)
            sheet_name = base_name[:31 - len(suffix)] + suffix
            k += 1

        ws = wb_out.create_sheet(title=sheet_name)

        # Escribir datos calculados (VALORES, no fórmulas) desde fila 3
        for i in range(df_calc.shape[0]):
            excel_r = 3 + i
            for j, h in enumerate(headers, start=1):
                ws.cell(row=excel_r, column=j, value=float(df_calc.iloc[i][h]))

        formatear_hoja(ws, desc_val, df_calc.shape[0])

    wb_out.save(salida_xlsx)
    return salida_xlsx

if __name__ == '__main__':
    salida = generar_hiley()
    print(salida)